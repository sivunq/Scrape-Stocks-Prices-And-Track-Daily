import urllib,requests,openpyxl
from openpyxl.styles import colors,Color, PatternFill, Font, Border,Side,Alignment
from openpyxl.cell import Cell
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime
from bs4 import BeautifulSoup

#Global variables
wb=openpyxl.load_workbook("FinTrack.xlsx")
sheet=wb["CurrentStocks"]
stocksRow=1
linksRow=2
dayChangeRow=3
shareCountRow=4
buyPriceRow=5
currentPriceRow=6
yesterdaysPriceRow=7
differenceRow=8
startRow=10
    
#read all constant values
def getDataFromSheet():
    startCol=2
    stocksData={
	"stockNames":[],
	"webLink":[],
	"buyPrice":[],
	"shareCount":[]
    }
    while(sheet.cell(row=1,column=startCol).value != "Total"):
        stocksData["stockNames"].append(sheet.cell(row=stocksRow,column=startCol).value)
        stocksData["webLink"].append(sheet.cell(row=linksRow,column=startCol).value)
        stocksData["shareCount"].append(sheet.cell(row=shareCountRow,column=startCol).value)
        stocksData["buyPrice"].append(sheet.cell(row=buyPriceRow,column=startCol).value)
        startCol=startCol+1
    return stocksData

#logic to find empty cell or update today's data
def getTodaysRow():
    flagSameDay=False
    todayRow=73 #Update
    todayDate=str(datetime.today().strftime('%d-%m-%Y'))

    while(True):
        if(str(sheet.cell(row=todayRow,column=1).value) == todayDate):
            flagSameDay=True
            break
        elif(str(sheet.cell(row=todayRow,column=1).value) == 'None'):
            sheet.cell(row=todayRow,column=1).value = todayDate
            break
        todayRow=todayRow+1
        
    return todayRow,flagSameDay

def main():
    stocksData=getDataFromSheet()
    todayRow,flagSameDay=getTodaysRow()

    total=0
    headers = {"user-agent" : "Mozilla/5.0 (Linux; Android 7.0; SM-G930V Build/NRD90M) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.125 Mobile Safari/537.36"}
    #iterate for each stock and scrape values from moneycontrol.com
    for index in range(1,len(stocksData["stockNames"])+4):
	#update stocks prices
        if(index>1 and index <= len(stocksData["stockNames"])+1):
            resp = requests.get(stocksData["webLink"][index-2], headers=headers)
            soup = BeautifulSoup(resp.content, "html.parser")
            mydivs = soup.findAll("div", {"class": "nsecp"})
            try:
                price=mydivs[0]["rel"]
            except Exception:
                price=1
                print("some issue happened...")
                
            price= round(float(price),2)
             
	    #update yesterdaysPrice
            if(not flagSameDay):
                sheet.cell(row=yesterdaysPriceRow,column=index).value=sheet.cell(row=currentPriceRow,column=index).value
            #dayChange
            yesterdaysPrice=sheet.cell(row=yesterdaysPriceRow,column=index).value
            sheet.cell(row=dayChangeRow,column=index).value=round((price-yesterdaysPrice)*stocksData["shareCount"][index-2],2)
            #currentPrice
            sheet.cell(row=currentPriceRow,column=index).value=price 
            #difference
            sheet.cell(row=differenceRow,column=index).value=price-stocksData["buyPrice"][index-2] 
            #profit
            profit=(float(price)-stocksData["buyPrice"][index-2])*stocksData["shareCount"][index-2]
            profit= round(profit,2)
            total=total+profit
            #fill profit
            sheet.cell(row=todayRow,column=index).value=profit
            print(index-1,".",stocksData["stockNames"][index-2],":",profit)
            #fill heat map
            colIndex=openpyxl.utils.cell.get_column_letter(index)
            sheet.conditional_formatting.add(colIndex+str(startRow)+":"+colIndex+str(todayRow) ,ColorScaleRule(start_type='min', start_value=0, start_color='F75E68',
					    mid_type='percentile', mid_value=50, mid_color='FFFA99',
					    end_type='max', end_value=100, end_color='36A460'))
		
	#style each column
        cellFill=sheet.cell(row=todayRow,column=index)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        cellFill.border = thin_border
        cellFill.alignment = Alignment(horizontal='center')

        darkGreen = openpyxl.styles.colors.Color(rgb='70AD47')
        myFill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=darkGreen)
        cellFill.fill = myFill

        index=index+1

    #update total and day change
    total=round(total,2)
    sheet.cell(row=todayRow,column=index-2).value=total

    prevTotal=sheet.cell(row=todayRow-1,column=index-2).value
    sheet.cell(row=todayRow,column=index-1).value=total-prevTotal

    wb.save("FinTrack.xlsx")
    wb.close()
    print("Done")
    input("Press Enter To Exit")

if __name__ == "__main__":
    main()
